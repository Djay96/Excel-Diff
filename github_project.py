import openpyxl
from openpyxl.styles import PatternFill, Font
from tqdm import tqdm
from PyQt5.QtWidgets import QApplication, QFileDialog, QInputDialog, QDialog, QListWidget
from PyQt5.QtCore import Qt
import os
import time
from datetime import timedelta
import sys
from PyQt5.QtWidgets import QVBoxLayout, QPushButton

def is_empty_row(row):
    try:
        return all(cell_value is None for cell_value in row)
    except Exception as e:
        print(f"An error occurred while checking if a row is empty: {str(e)}. Please check your Excel files.")

def get_trade_rows(sheet, key_column, column_mapping):
    try:
        trade_rows = {}
        header_row = list(sheet.iter_rows(min_row=1, max_row=1, values_only=True))[0]
        key_column_index = header_row.index(key_column)
        if key_column_index is None:
            return trade_rows, None, "Key column not found. Please make sure the selected key column exists in your Excel file."
        for row in sheet.iter_rows(min_row=2, values_only=True):
            key_value = row[key_column_index]
            trade_rows[key_value] = row
        return trade_rows, column_mapping, None
    except Exception as e:
        return None, None, f"An error occurred while processing the Excel sheet: {str(e)}. Please check your Excel file."

def rows_are_different(row1, row2, column_mapping):
    try:
        if row1 is None or row2 is None:
            return True
        for col_name in column_mapping:
            if row1[column_mapping[col_name]] != row2[column_mapping[col_name]]:
                return True
        return False
    except Exception as e:
        print(f"An error occurred while comparing rows: {str(e)}. Please check your Excel files.")

def write_row_to_output(sheet_out, output_row_num, output_column_start, row, other_row, is_row1, file1, file2, column_mapping, diff_columns):
    try:
        status_cell=sheet_out.cell(row=output_row_num,column=1)
        file1_basename = os.path.basename(file1)
        file2_basename = os.path.basename(file2)
        if other_row is None:
            if is_row1:
                status_cell.value=f"Missing from {file2_basename if file1_basename != file2_basename else 'file2'}"
            else:
                status_cell.value=f"Missing from {file1_basename if file1_basename != file2_basename else 'file1'}"
            trade_id_cell=sheet_out.cell(row=output_row_num,column=8)
            trade_id_cell.fill=PatternFill(fill_type="solid",fgColor="FFFF00")
        else:
            status_cell.value=f"Differences {file1_basename if is_row1 else file2_basename}"
        for col_name in column_mapping:
            col_num = column_mapping[col_name] + output_column_start
            cell=sheet_out.cell(row=output_row_num,column=col_num)
            cell.value=row[column_mapping[col_name]]
            if other_row is not None and len(other_row) > column_mapping[col_name] and cell.value != other_row[column_mapping[col_name]]:
                cell.fill=PatternFill(fill_type="solid",fgColor="FFFF00")
                diff_columns.add(col_name)  # Add the column name to the set of columns with differences
    except Exception as e:
        print(f"An error occurred while writing the output: {str(e)}. Please check your Excel files.")

class FileDialog(QDialog):
    def __init__(self, *args, **kwargs):
        super(FileDialog, self).__init__(*args, **kwargs)
        self.setWindowFlags(self.windowFlags() | Qt.WindowStaysOnTopHint)

def get_file_path(caption):
    try:
        dialog = FileDialog()
        file_path, _ = QFileDialog.getOpenFileName(dialog, caption, "", "Excel files (*.xlsx)")
        return file_path
    except Exception as e:
        print(f"An error occurred while opening the file dialog: {str(e)}. Please try again.")

def get_key_column(header_row):
    dialog = QDialog()
    dialog.setWindowFlags(dialog.windowFlags() | Qt.WindowStaysOnTopHint)
    selected_column, ok = QInputDialog.getItem(dialog, "Select Key Column", "Choose the key column for comparison:", header_row, 0, False)
    return selected_column if ok else None

class MultiColumnDialog(QDialog):
    def __init__(self, columns, *args, **kwargs):
        super(MultiColumnDialog, self).__init__(*args, **kwargs)
        self.setWindowTitle("Select Columns to Exclude")
        self.setWindowFlags(self.windowFlags() | Qt.WindowStaysOnTopHint)
        self.columns = columns
        self.selected_columns = []

    def select_columns(self):
        list_widget = QListWidget(self)
        list_widget.addItems(["<None>"] + self.columns)
        list_widget.setSelectionMode(QListWidget.MultiSelection)
        
        # OK button
        ok_button = QPushButton('OK', self)
        ok_button.clicked.connect(self.accept)

        # Cancel button
        cancel_button = QPushButton('Cancel', self)
        cancel_button.clicked.connect(self.reject)

        # Layout
        layout = QVBoxLayout()
        layout.addWidget(list_widget)
        layout.addWidget(ok_button)
        layout.addWidget(cancel_button)
        self.setLayout(layout)

        # Set size
        self.resize(600, 400)

        result = self.exec_()
        if result == QDialog.Accepted:
            self.selected_columns = [item.text() for item in list_widget.selectedItems() if item.text() != "<None>"]

def get_excluded_columns(header_row):
    try:
        dialog = MultiColumnDialog(list(header_row))  # Convert header_row to a list
        dialog.select_columns()
        return dialog.selected_columns
    except Exception as e:
        print(f"An error occurred while selecting the excluded columns: {str(e)}. Please try again.")

def compare_excel_files():
    try:
        app = QApplication(sys.argv)
        start_time = time.time()
        with tqdm(total=6) as pbar:
            file1 = get_file_path("Select the first Excel file")
            if not file1:
                print("No file was selected. Please select an Excel file.")
                return
            print(f"Input file 1: {file1}")
            file2 = get_file_path("Select the second Excel file")
            if not file2:
                print("No file was selected. Please select an Excel file.")
                return
            print(f"Input file 2: {file2}")
            output_file = "Diff_Output.xlsx"
            pbar.update(1)
            wb1 = openpyxl.load_workbook(file1)
            wb2 = openpyxl.load_workbook(file2)
            pbar.update(1)
            if len(wb1.sheetnames) > 1 or len(wb2.sheetnames) > 1:
                wb_out = openpyxl.Workbook()
                sheet_out = wb_out.active
                sheet_out.cell(row=1, column=1).value = "Error: More than one sheet found in the input files. Please make sure each file has only one sheet."
                wb_out.save(output_file)
                return

            sheet1 = wb1.active
            sheet2 = wb2.active
            header_row = list(sheet1.iter_rows(min_row=1, max_row=1, values_only=True))[0]
            key_column = get_key_column(header_row)
            if not key_column:
                print("No key column was selected. Please select a key column.")
                return

            excluded_columns = get_excluded_columns(header_row)
            print(f"Excluded columns: {excluded_columns}")

            # Filter column mapping to exclude selected columns
            column_mapping = {col_name: i for i, col_name in enumerate(header_row) if col_name not in excluded_columns}

            trade_rows1, _, error1 = get_trade_rows(sheet1, key_column, column_mapping)
            trade_rows2, _, error2 = get_trade_rows(sheet2, key_column, column_mapping)

            
            pbar.update(1)
            if error1 or error2:
                wb_out = openpyxl.Workbook()
                sheet_out = wb_out.active
                sheet_out.cell(row=1, column=1).value = error1 if error1 else error2
                wb_out.save(output_file)
                return
            all_trade_ids = set(list(trade_rows1.keys()) + list(trade_rows2.keys()))
            pbar.update(1)
            wb_out = openpyxl.Workbook()
            sheet_out = wb_out.active
            output_row_num = 1
            sheet_out.cell(row=output_row_num, column=1).value = "Status"

            # Create the common column mapping, excluding the specified columns
            common_column_mapping = {
                col_name: column_mapping[col_name] 
                for col_name in column_mapping
                if col_name in column_mapping and col_name not in excluded_columns
            }

            header_row = [col_name for col_name in common_column_mapping]
            for col_num, cell_value in enumerate(header_row, start=2):
                cell = sheet_out.cell(row=output_row_num, column=col_num)
                cell.value = cell_value
                # Make the header cells bold
                cell.font = Font(bold=True)
            output_row_num += 1
            diff_columns = set()
            for trade_id in all_trade_ids:
                row1 = trade_rows1.get(trade_id, None)
                row2 = trade_rows2.get(trade_id, None)
                if row1 is None or row2 is None or rows_are_different(row1,row2, common_column_mapping):
                    if row1 is not None and not is_empty_row(row1):
                        write_row_to_output(sheet_out,output_row_num, 2,row1,row2,True, file1, file2, common_column_mapping, diff_columns)
                        # Make the first cell of the row bold
                        sheet_out.cell(row=output_row_num, column=1).font = Font(bold=True)
                        output_row_num += 1
                    if row2 is not None and not is_empty_row(row2):
                        write_row_to_output(sheet_out,output_row_num, 2,row2,row1,False, file1, file2, common_column_mapping, diff_columns)
                        # Make the first cell of the row bold
                        sheet_out.cell(row=output_row_num, column=1).font = Font(bold=True)
                        output_row_num += 1
                pbar.update(0.5/len(all_trade_ids))
            pbar.update(0.5)
            # Highlight the column headers with differences at the very end
            for col_name in diff_columns:
                col_num = common_column_mapping[col_name] + 2
                cell=sheet_out.cell(row=1,column=col_num)
                cell.fill=PatternFill(fill_type="solid",fgColor="FFFF00")
            # Save the workbook after all operations (including highlighting) have been done
            wb_out.save(output_file)
            elapsed_time = time.time() - start_time
            formatted_time = str(timedelta(seconds=int(elapsed_time)))
            full_output_path = os.path.abspath(output_file)
            print(f"Comparison done. Differences written to: {full_output_path}. Total elapsed time: {formatted_time}.")
    except Exception as e:
        print(f"An unexpected error occurred: {str(e)}. Please check your Excel files and try again.")

compare_excel_files()


